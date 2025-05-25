import pandas as pd
import os
from openpyxl import load_workbook

def create_movimento_sheet():
    """Cria a planilha de movimento com estrutura simplificada."""
    # Estrutura para 2024 (todos os meses)
    data_2024 = {
        'Empresa': [],
        'Código': [],
        'Classificação': [],
        'Descrição': [],
        'Jan/2024': [],
        'Fev/2024': [],
        'Mar/2024': [],
        'Abr/2024': [],
        'Mai/2024': [],
        'Jun/2024': [],
        'Jul/2024': [],
        'Ago/2024': [],
        'Set/2024': [],
        'Out/2024': [],
        'Nov/2024': [],
        'Dez/2024': [],
        'Saldo Acumulado': []
    }
    
    # Estrutura para 2025 (trimestral)
    data_2025 = {
        'Empresa': [],
        'Código': [],
        'Classificação': [],
        'Descrição': [],
        'Jan/2025': [],
        'Fev/2025': [],
        'Mar/2025': [],
        'Saldo Acumulado': []
    }
    
    # Cria DataFrames vazios
    df_2024 = pd.DataFrame(data_2024)
    df_2025 = pd.DataFrame(data_2025)
    
    return df_2024, df_2025

def create_balancete_structure():
    """Cria a estrutura base do balancete."""
    return [
        'ATIVO',
        'PASSIVO',
        'PATRIMONIO LIQUIDO',
        'ESTOQUES',
        'COMPENSACOES',
        'RECEITA OPERACIONAL BRUTA',
        'DEDUCOES/CUSTOS/DESPESAS',
        'APURACAO DO RESULTADO',
        '',  # Linha em branco
        'CONTAS DEVEDORAS',
        'CONTAS CREDORAS',
        '',  # Linha em branco
        'RESULTADO DO MES',
        'RESULTADO DO EXERCÍCIO'
    ]

def create_consolidated_sheet():
    """Cria a planilha consolidada com todas as empresas."""
    # Estrutura base
    data = {
        'Empresa': [],
        'Info': [],
        '01/2025': [],
        '02/2025': [],
        '03/2025': [],
        'Saldo Acumulado': []
    }
    
    return pd.DataFrame(data)

def create_company_sheet(year):
    """Cria a planilha para uma empresa específica em um ano."""
    # Estrutura base
    if year == 2024:
        data = {
            'Info': [],
            'Jan/2024': [],
            'Fev/2024': [],
            'Mar/2024': [],
            'Abr/2024': [],
            'Mai/2024': [],
            'Jun/2024': [],
            'Jul/2024': [],
            'Ago/2024': [],
            'Set/2024': [],
            'Out/2024': [],
            'Nov/2024': [],
            'Dez/2024': [],
            'Saldo Consolidado': []
        }
    else:  # 2025 (trimestral)
        data = {
            'Info': [],
            'Jan/2025': [],
            'Fev/2025': [],
            'Mar/2025': [],
            'Saldo Consolidado': []
        }
    
    return pd.DataFrame(data)

def save_to_excel(df, filename, sheet_name):
    """Salva o DataFrame em um arquivo Excel com formatação."""
    if os.path.exists(filename):
        # Carrega o workbook existente
        book = load_workbook(filename)
        
        # Remove a aba se ela já existir
        if sheet_name in book.sheetnames:
            idx = book.sheetnames.index(sheet_name)
            book.remove(book.worksheets[idx])
        
        # Salva o workbook
        book.save(filename)
        
        # Salva o DataFrame na nova aba
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Ajusta o formato das colunas
            worksheet = writer.sheets[sheet_name]
            worksheet.column_dimensions['A'].width = 40  # Info
            for col in worksheet.columns:
                if col[0].column_letter != 'A':  # Colunas de valores
                    worksheet.column_dimensions[col[0].column_letter].width = 15
    else:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Ajusta o formato das colunas
            worksheet = writer.sheets[sheet_name]
            worksheet.column_dimensions['A'].width = 40  # Info
            for col in worksheet.columns:
                if col[0].column_letter != 'A':  # Colunas de valores
                    worksheet.column_dimensions[col[0].column_letter].width = 15

def create_sheet_2025():
    """Cria a estrutura para 2025."""
    # Estrutura base
    data = {
        'Info': [],
        '01/2025': [],
        '02/2025': [],
        '03/2025': [],
        'Saldo Acumulado': []
    }
    return pd.DataFrame(data)

def create_sheet_2024():
    """Cria a estrutura para 2024."""
    # Estrutura base
    data = {
        'Info': [],
        '01/2024': [],
        '02/2024': [],
        '03/2024': [],
        '04/2024': [],
        '05/2024': [],
        '06/2024': [],
        '07/2024': [],
        '08/2024': [],
        '09/2024': [],
        '10/2024': [],
        '11/2024': [],
        '12/2024': [],
        'Saldo Acumulado': []
    }
    return pd.DataFrame(data)

def create_empresa_files():
    """Cria os arquivos base para cada empresa."""
    # Lista de empresas
    empresas = {
        'porto_santa_maria': 'PORTO SANTA MARIA EMPREENDIMENTOS TURIST',
        'marina': 'MARINA',
        'porto_marina': 'PORTO MARINA'
    }
    
    # Cria a estrutura do balancete
    balancete_info = create_balancete_structure()
    
    # Para cada empresa
    for filename, empresa in empresas.items():
        # Cria o DataFrame para 2024
        df_2024 = create_sheet_2024()
        df_2024['Info'] = balancete_info
        
        # Cria o DataFrame para 2025
        df_2025 = create_sheet_2025()
        df_2025['Info'] = balancete_info
        
        # Salva os DataFrames no arquivo da empresa
        excel_file = f'{filename}.xlsx'
        if os.path.exists(excel_file):
            os.remove(excel_file)
            
        save_to_excel(df_2024, excel_file, '2024')
        save_to_excel(df_2025, excel_file, '2025')
        
        print(f"Arquivo {excel_file} criado com sucesso!")

def main():
    create_empresa_files()
    print("Todos os arquivos foram criados com sucesso!")

if __name__ == "__main__":
    main() 